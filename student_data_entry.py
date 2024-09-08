import os
import pandas as pd
import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from tkcalendar import DateEntry
from PIL import Image, ImageTk

class StudentDataApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Job Oriented Training Programme Students & Data Entry")

        # Set window background color
        self.root.configure(bg='#f0f0f0')

        # Create a canvas and a scrollbar
        self.canvas = tk.Canvas(root, bg='#f0f0f0')
        self.scrollbar = tk.Scrollbar(root, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = tk.Frame(self.canvas, bg='#f0f0f0')

        # Pack the widgets
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )

        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        # Load and place logo
        self.logo_image = Image.open("kanishka_logo.png")
        self.logo_image = self.logo_image.resize((200, 100), Image.Resampling.LANCZOS)
        self.logo_photo = ImageTk.PhotoImage(self.logo_image)
        self.logo_label = tk.Label(self.scrollable_frame, image=self.logo_photo, bg='#f0f0f0')
        self.logo_label.grid(row=0, column=0, columnspan=2, pady=10)

        # Label and Entry widget configuration
        label_font = ("Arial", 12, "bold")
        entry_font = ("Arial", 12)
        entry_width = 30
        label_bg = '#f0f0f0'
        label_fg = '#333333'

        # Helper function to create labels and entries
        def create_label_entry(row, text, column=0):
            label = tk.Label(self.scrollable_frame, text=text, font=label_font, bg=label_bg, fg=label_fg)
            label.grid(row=row, column=column, sticky='e', padx=10, pady=5)
            entry = tk.Entry(self.scrollable_frame, font=entry_font, width=entry_width)
            entry.grid(row=row, column=column + 1, pady=5)
            entry.bind("<Return>", self.focus_next_widget)
            return entry

        # Helper function to create dropdown menus
        def create_dropdown(row, text, options, column=0):
            label = tk.Label(self.scrollable_frame, text=text, font=label_font, bg=label_bg, fg=label_fg)
            label.grid(row=row, column=column, sticky='e', padx=10, pady=5)
            var = tk.StringVar()
            dropdown = tk.OptionMenu(self.scrollable_frame, var, *options)
            dropdown.config(font=entry_font, width=entry_width - 5)
            dropdown.grid(row=row, column=column + 1, pady=5)
            var.set(options[0])
            return var
        
        # Section Titles
        section_title_font = ("Arial", 14, "bold")
        section_title_fg = 'red'
        section_title_bg = 'red'

        # Additional Fields Section
        tk.Label(self.scrollable_frame, text="Personal Information", font=section_title_font, fg='red', bg='#f0f0f0').grid(row=24, column=0, columnspan=2, pady=10)        
        self.full_name_entry = create_label_entry(25, "Name In Full")
        self.initial_name_entry = create_label_entry(26, "Name With Initial")
        self.dob_label = tk.Label(self.scrollable_frame, text="Date Of Birth", font=label_font, bg=label_bg, fg=label_fg)
        self.dob_label.grid(row=28, column=0, sticky='e', padx=10, pady=5)
        self.dob_entry = DateEntry(self.scrollable_frame, width=entry_width - 3, font=entry_font)
        self.dob_entry.grid(row=28, column=1, pady=5)
        self.dob_entry.bind("<Return>", self.focus_next_widget)

        self.nic_entry = create_label_entry(28, "National Identity Card No")
        
        tk.Label(self.scrollable_frame, text="Gender", font=label_font, bg=label_bg, fg=label_fg).grid(row=29, column=0, sticky='e', padx=10, pady=5)
        self.gender_var = tk.StringVar()
        self.gender_dropdown = tk.OptionMenu(self.scrollable_frame, self.gender_var, "Male", "Female", "Rather Not Say")
        self.gender_dropdown.config(font=entry_font, width=entry_width - 5)
        self.gender_dropdown.grid(row=29, column=1, pady=5)
        self.gender_var.set("Select Gender")
        
        self.postal_address_entry = create_label_entry(30, "Postal Address")
        self.id_address_entry = create_label_entry(31, "Address In ID")
        self.permanent_address_entry = create_label_entry(32, "Permanent Address")
        self.current_address_entry = create_label_entry(33, "Current Address")
        self.mobile_number_entry = create_label_entry(34, "Mobile Number")
        self.land_phone_number_entry = create_label_entry(35, "Land Phone Number")
        self.whatsapp_number_entry = create_label_entry(36, "Whatsapp Number")
        self.facebook_url_entry = create_label_entry(37, "Facebook URL")
        self.instagram_id_entry = create_label_entry(38, "Instagram ID")
        self.email_entry = create_label_entry(39, "Email")

        # Geographical Details Section
        tk.Label(self.scrollable_frame, text="Geographical Details", font=section_title_font, fg=section_title_fg).grid(row=40, column=0, columnspan=2, pady=10)
        
        provinces = ["Select Province", "Central Province", "Eastern Province", "Northern Province", "Southern Province", "Western Province", "North Western Province", "North Central Province", "Uva Province", "Sabaragamuwa Province"]
        self.province_var = create_dropdown(41, "Province", provinces)
        districts = [
            "Select District", "Colombo", "Gampaha", "Kalutara", "Kandy", "Matale", "Nuwara Eliya", 
            "Galle", "Matara", "Hambantota", "Jaffna", "Kilinochchi", "Mannar", "Vavuniya", 
            "Mullaitivu", "Batticaloa", "Ampara", "Trincomalee", "Kurunegala", "Puttalam", 
            "Anuradhapura", "Polonnaruwa", "Badulla", "Moneragala", "Ratnapura", "Kegalle"
        ]
        self.district_var = create_dropdown(42, "District", districts)

        ds_divisions = [
            "Select DS Division", "Alawwa", "Ambanpola", "Bamunakotuwa", "Bingiriya", "Polgahawela", 
            "Polpithigama", "Pothuhera", "Ridigama", "Wariyapola", "Weerambugedara", "Kobeigane", 
            "Kuliyapitiya East", "Kuliyapitiya West", "Kurunegala", "Mahawa", "Mallawapitiya", 
            "Maspotha", "Mawathagama", "Narammala", "Panduwasnuwara East", "Panduwasnuwara West", 
            "Udubaddawa"
        ]
        self.ds_division_var = create_dropdown(43, "DS Division", ds_divisions)
        self.gn_division_entry = create_label_entry(44, "GN Division")
        self.nearest_city_entry = create_label_entry(45, "Nearest City")

        # Parent/Guardian Information Section
        tk.Label(self.scrollable_frame, text="Parent or Guardian Information", font=section_title_font, fg=section_title_fg).grid(row=46, column=0, columnspan=2, pady=10)
        
        self.guardian_name_entry = create_label_entry(47, "Full Name")
        self.guardian_dob_label = tk.Label(self.scrollable_frame, text="Date of Birth", font=label_font, bg=label_bg, fg=label_fg)
        self.guardian_dob_label.grid(row=48, column=0, sticky='e', padx=10, pady=5)
        self.guardian_dob_entry = DateEntry(self.scrollable_frame, width=entry_width - 3, font=entry_font)
        self.guardian_dob_entry.grid(row=48, column=1, pady=5)
        self.guardian_dob_entry.bind("<Return>", self.focus_next_widget)
        self.guardian_age_entry = create_label_entry(49, "Age")
        self.guardian_occupation_entry = create_label_entry(50, "Occupation")
        self.guardian_contact_number_entry = create_label_entry(51, "Contact Number")
        self.guardian_whatsapp_number_entry = create_label_entry(52, "WhatsApp Number")
        self.guardian_address_entry = create_label_entry(53, "Address")
        self.guardian_email_entry = create_label_entry(54, "Email")

        # Educational Qualifications Section
        tk.Label(self.scrollable_frame, text="Educational Qualifications", font=section_title_font, fg=section_title_fg).grid(row=55, column=0, columnspan=2, pady=10)
        
        # G.C.E (O/L)
        tk.Label(self.scrollable_frame, text="G.C.E (O/L) Examination", font=label_font, bg=label_bg, fg=label_fg).grid(row=56, column=0, columnspan=2, pady=5)
        self.ol_school_entry = create_label_entry(57, "Name of School/College")
        self.ol_admission_number_entry = create_label_entry(58, "Admission Number of Examination (O/L)")
        self.ol_year_label = tk.Label(self.scrollable_frame, text="Year of Sitting", font=label_font, bg=label_bg, fg=label_fg)
        self.ol_year_label.grid(row=59, column=0, sticky='e', padx=10, pady=5)
        self.ol_year_entry = DateEntry(self.scrollable_frame, width=entry_width - 3, font=entry_font)
        self.ol_year_entry.grid(row=59, column=1, pady=5)
        self.ol_year_entry.bind("<Return>", self.focus_next_widget)
        self.ol_award_achievements_entry = create_label_entry(60, "Award Achievements")

        # G.C.E (A/L)
        tk.Label(self.scrollable_frame, text="G.C.E (A/L) Examination", font=label_font, bg=label_bg, fg=label_fg).grid(row=61, column=0, columnspan=2, pady=5)
        self.al_school_entry = create_label_entry(62, "Name of School/College")
        self.al_admission_number_entry = create_label_entry(63, "Admission Number of Examination (A/L)")
        self.al_year_label = tk.Label(self.scrollable_frame, text="Year of Sitting", font=label_font, bg=label_bg, fg=label_fg)
        self.al_year_label.grid(row=64, column=0, sticky='e', padx=10, pady=5)
        self.al_year_entry = DateEntry(self.scrollable_frame, width=entry_width - 3, font=entry_font)
        self.al_year_entry.grid(row=64, column=1, pady=5)
        self.al_year_entry.bind("<Return>", self.focus_next_widget)
        self.al_award_achievements_entry = create_label_entry(65, "Award Achievements")

        # Professional Qualifications Section
        tk.Label(self.scrollable_frame, text="Professional Qualifications", font=section_title_font, fg=section_title_fg).grid(row=66, column=0, columnspan=2, pady=10)
        
        self.pro_institution_entry = create_label_entry(67, "Institution")
        self.pro_result_achievement_entry = create_label_entry(68, "Result Achievement")
        self.pro_pass_out_year_label = tk.Label(self.scrollable_frame, text="Pass Out Year", font=label_font, bg=label_bg, fg=label_fg)
        self.pro_pass_out_year_label.grid(row=69, column=0, sticky='e', padx=10, pady=5)
        self.pro_pass_out_year_entry = DateEntry(self.scrollable_frame, width=entry_width - 3, font=entry_font)
        self.pro_pass_out_year_entry.grid(row=69, column=1, pady=5)
        self.pro_pass_out_year_entry.bind("<Return>", self.focus_next_widget)
        self.pro_areas_covered_entry = create_label_entry(70, "Areas Covered")
        self.pro_any_other_entry = create_label_entry(71, "Any Other")

        # Other Available Qualifications Section
        tk.Label(self.scrollable_frame, text="Other Available Qualifications", font=section_title_font, fg=section_title_fg).grid(row=72, column=0, columnspan=2, pady=10)
        
        self.other_qualifications_entry = create_label_entry(73, "Other Qualifications")

        # Special Talents Section
        tk.Label(self.scrollable_frame, text="Special Talents", font=section_title_font, fg=section_title_fg).grid(row=74, column=0, columnspan=2, pady=10)
        
        self.special_talent_entry = create_label_entry(75, "Special Talent (e.g., Dancing, Singing, Writing, etc.)")

        # Employment Status Section
        tk.Label(self.scrollable_frame, text="Employment Status", font=section_title_font, fg=section_title_fg).grid(row=76, column=0, columnspan=2, pady=10)
        
        self.employment_institution_entry = create_label_entry(77, "Name of the Institution")
        self.employment_job_title_entry = create_label_entry(78, "Job Title")
        self.employment_joining_year_label = tk.Label(self.scrollable_frame, text="Joining Year", font=label_font, bg=label_bg, fg=label_fg)
        self.employment_joining_year_label.grid(row=79, column=0, sticky='e', padx=10, pady=5)
        self.employment_joining_year_entry = DateEntry(self.scrollable_frame, width=entry_width - 3, font=entry_font)
        self.employment_joining_year_entry.grid(row=79, column=1, pady=5)
        self.employment_joining_year_entry.bind("<Return>", self.focus_next_widget)
        self.employment_year_left_label = tk.Label(self.scrollable_frame, text="Year Left", font=label_font, bg=label_bg, fg=label_fg)
        self.employment_year_left_label.grid(row=80, column=0, sticky='e', padx=10, pady=5)
        self.employment_year_left_entry = DateEntry(self.scrollable_frame, width=entry_width - 3, font=entry_font)
        self.employment_year_left_entry.grid(row=80, column=1, pady=5)
        self.employment_year_left_entry.bind("<Return>", self.focus_next_widget)

        # Student Information Section
        tk.Label(self.scrollable_frame, text="Student & Modules & Lectures Information", font=section_title_font, bg=section_title_bg, fg=section_title_fg).grid(row=81, column=0, columnspan=2, pady=10)

        self.student_no_entry = create_label_entry(82, "Student No")
        self.student_name_entry = create_label_entry(83, "Student Name")
        self.program_name_entry = create_label_entry(84, "Program Name")
        self.jot_activity1_entry = create_label_entry(85, "JOT Activity")
        self.phicycle_activity1_entry = create_label_entry(86, "Phicycle Activity")
        self.lab_section1_entry = create_label_entry(87, "Lab Section")
        self.assignment1_entry = create_label_entry(88, "Assignment")
        self.assessment1_entry = create_label_entry(89, "Assessment")
        self.practical1_entry = create_label_entry(90, "Practical")
        self.module1_title_entry = create_label_entry(91, "Module Title")
        self.module1_subject1_entry = create_label_entry(92, "Module Subject")

# Module 1 Hours and Minutes
        self.module1_hours_label = tk.Label(self.scrollable_frame, text="Module Hours", font=label_font, bg=label_bg, fg=label_fg)
        self.module1_hours_label.grid(row=93, column=0, sticky='e', padx=10, pady=5)
        self.module1_hours_spinbox = tk.Spinbox(self.scrollable_frame, from_=0, to=23, wrap=True, state="readonly", width=5, font=entry_font)
        self.module1_hours_spinbox.grid(row=93, column=1, sticky='w', pady=5)
        self.module1_hours_spinbox.bind("<Return>", self.focus_next_widget)
        self.module1_minutes_spinbox = tk.Spinbox(self.scrollable_frame, from_=0, to=59, wrap=True, state="readonly", width=5, font=entry_font)
        self.module1_minutes_spinbox.grid(row=93, column=1, sticky='e', pady=5)
        self.module1_minutes_spinbox.bind("<Return>", self.focus_next_widget)

        self.module1_supervisor_entry = create_label_entry(94, "Module Supervisor")
        self.module1_lecturer_entry = create_label_entry(95, "Module Lecturer")
        self.module1_satisfaction_entry = create_label_entry(96, "Lecturer or Supervisor Satisfaction")
        self.module1_lecturer_no_entry = create_label_entry(97, "Module Lecturer No")
        self.module1_supervisor_no_entry = create_label_entry(98, "Module Supervisor No")
        self.module1_title_no_entry = create_label_entry(99, "Module Title No")
        self.module1_subject_no_entry = create_label_entry(100, "Module Subject No")

        # Subject Hours and Minutes
        self.subject_hours_label = tk.Label(self.scrollable_frame, text="Subject Hours", font=label_font, bg=label_bg, fg=label_fg)
        self.subject_hours_label.grid(row=101, column=0, sticky='e', padx=10, pady=5)
        self.subject_hours_spinbox = tk.Spinbox(self.scrollable_frame, from_=0, to=23, wrap=True, state="readonly", width=5, font=entry_font)
        self.subject_hours_spinbox.grid(row=101, column=1, sticky='w', pady=5)
        self.subject_hours_spinbox.bind("<Return>", self.focus_next_widget)
        self.subject_minutes_spinbox = tk.Spinbox(self.scrollable_frame, from_=0, to=59, wrap=True, state="readonly", width=5, font=entry_font)
        self.subject_minutes_spinbox.grid(row=101, column=1, sticky='e', pady=5)
        self.subject_minutes_spinbox.bind("<Return>", self.focus_next_widget)

        self.module_date_label = tk.Label(self.scrollable_frame, text="Module Date", font=label_font, bg=label_bg, fg=label_fg)
        self.module_date_label.grid(row=102, column=0, sticky='e', padx=10, pady=5)
        self.module_date_entry = DateEntry(self.scrollable_frame, font=entry_font)
        self.module_date_entry.grid(row=102, column=1, pady=5)
        self.module_date_entry.bind("<Return>", self.focus_next_widget)

        self.subject_date_label = tk.Label(self.scrollable_frame, text="Subject Date", font=label_font, bg=label_bg, fg=label_fg)
        self.subject_date_label.grid(row=103, column=0, sticky='e', padx=10, pady=5)
        self.subject_date_entry = DateEntry(self.scrollable_frame, font=entry_font)
        self.subject_date_entry.grid(row=103, column=1, pady=5)
        self.subject_date_entry.bind("<Return>", self.focus_next_widget)

        self.final_approval_entry = create_label_entry(104, "Final Approval Note")
 

        # Declaration Section
        tk.Label(self.scrollable_frame, text="Declaration", font=section_title_font, fg=section_title_fg).grid(row=105, column=0, columnspan=2, pady=10)
        
        declaration_text = "I, declare that the information provided herein is accurate, complete, and true to my knowledge. " \
                           "I understand that any false statement or omission may have legal implications and result in " \
                           "the denial or revocation of services, employment, or legal standings dependent on this declaration. " \
                           "By signing this declaration, I consent and agree to the above-mentioned terms concerning my biodata."
        
        self.declaration_label = tk.Label(self.scrollable_frame, text=declaration_text, font=("Arial", 10), bg=label_bg, fg=label_fg, wraplength=600, justify="left")
        self.declaration_label.grid(row=105, column=0, columnspan=2, padx=10, pady=10)
        
        self.agree_var = tk.IntVar()
        self.agree_checkbox = tk.Checkbutton(self.scrollable_frame, text="I Agree", variable=self.agree_var, font=entry_font, bg=label_bg)
        self.agree_checkbox.grid(row=106, column=0, columnspan=2, pady=5)
        # ... (previous code remains the same)

        # Phase 1 Section
        tk.Label(self.scrollable_frame, text="Phase 1", font=section_title_font, fg=section_title_fg).grid(row=109, column=0, columnspan=2, pady=10)

        self.se_no_entry = create_label_entry(110, "SE No")
        self.in_no_entry = create_label_entry(111, "IN No")
        self.phase1_name_entry = create_label_entry(112, "Name With Initial")
        self.phase1_age_entry = create_label_entry(113, "Age")
        self.phase1_phone_entry = create_label_entry(114, "Phone No")
        self.phase1_whatsapp_entry = create_label_entry(115, "Whatsapp No")
        self.phase1_email_entry = create_label_entry(116, "Email")
        self.phase1_address_entry = create_label_entry(117, "Address")
        self.phase1_ol_result_entry = create_label_entry(118, "O/L Result")

        note_options = ["Normal", "Cool"]
        self.phase1_note_var = create_dropdown(119, "Note", note_options)

        ol_sheet_options = ["Yes", "No"]
        self.phase1_ol_sheet_var = create_dropdown(120, "O/L Sheet", ol_sheet_options)

        self.phase1_birth_cert_entry = create_label_entry(121, "Birth Certificate")
        self.phase1_id_copy_entry = create_label_entry(122, "ID Copy")
        self.phase1_gs_copy_entry = create_label_entry(123, "GS Copy")
        self.phase1_edu_cert_entry = create_label_entry(124, "Education Certificate")

        # Phase 2 Section
        tk.Label(self.scrollable_frame, text="Phase 2", font=section_title_font, fg=section_title_fg).grid(row=125, column=0, columnspan=2, pady=10)

        self.phase2_serial_no_entry = create_label_entry(126, "Serial No")
        self.phase2_interview_no_entry = create_label_entry(127, "Interview No")
        self.phase2_name_entry = create_label_entry(128, "Name of the Applicant")
        self.phase2_height_entry = create_label_entry(129, "Height")
        self.phase2_weight_entry = create_label_entry(130, "Weight")
        self.phase2_blood_group_entry = create_label_entry(131, "Blood Group")

        dress_code_options = ["Ok", "SomeWhat", "Cool"]
        self.phase2_dress_code_var = create_dropdown(132, "Dress Code", dress_code_options)

        appearance_options = ["Good", "Normal", "Tensionable"]
        self.phase2_appearance_var = create_dropdown(133, "Appearance", appearance_options)

        course_understanding_options = ["Excellent", "Somewhat", "Not"]
        self.phase2_course_understanding_var = create_dropdown(134, "Course Understanding", course_understanding_options)

        self.phase2_age_entry = create_label_entry(135, "Age")
        self.phase2_location_entry = create_label_entry(136, "Location")
        self.phase2_distance_entry = create_label_entry(137, "Distance Up to Giriulla")

        status_options = ["Suitable", "Somewhat", "Normal"]
        self.phase2_status_var = create_dropdown(138, "Status", status_options)

        # Phase 3 Section
        tk.Label(self.scrollable_frame, text="Phase 3", font=section_title_font, fg=section_title_fg).grid(row=139, column=0, columnspan=2, pady=10)

        self.phase3_interview_no_entry = create_label_entry(140, "Interview No")

        # Education Sub-section
        tk.Label(self.scrollable_frame, text="Education", font=("Arial", 12, "bold"), bg='#f0f0f0', fg='#333333').grid(row=141, column=0, columnspan=2, pady=5)
        self.phase3_gce_ol_entry = create_label_entry(142, "GCE O/L")
        self.phase3_gce_al_entry = create_label_entry(143, "GCE A/L")
        self.phase3_higher_study_entry = create_label_entry(144, "Higher Study")

        # Skills Sub-section
        tk.Label(self.scrollable_frame, text="Skills", font=("Arial", 12, "bold"), bg='#f0f0f0', fg='#333333').grid(row=145, column=0, columnspan=2, pady=5)
        self.phase3_train_by_entry = create_label_entry(146, "TrainBy")
        self.phase3_self_entry = create_label_entry(147, "Self")

        # Vocational Qualification Sub-section
        tk.Label(self.scrollable_frame, text="Vocational Qualification", font=("Arial", 12, "bold"), bg='#f0f0f0', fg='#333333').grid(row=148, column=0, columnspan=2, pady=5)
        self.phase3_voc_category_entry = create_label_entry(149, "Category")
        self.phase3_voc_training_period_entry = create_label_entry(150, "Training Period")
        self.phase3_voc_experiences_entry = create_label_entry(151, "Experiences")

        # Professional Qualification Sub-section
        tk.Label(self.scrollable_frame, text="Professional Qualification", font=("Arial", 12, "bold"), bg='#f0f0f0', fg='#333333').grid(row=152, column=0, columnspan=2, pady=5)
        self.phase3_prof_category_entry = create_label_entry(153, "Category")
        self.phase3_prof_institute_entry = create_label_entry(154, "Institute")
        self.phase3_prof_experience_entry = create_label_entry(155, "Experience")

        # Others Sub-section
        tk.Label(self.scrollable_frame, text="Others", font=("Arial", 12, "bold"), bg='#f0f0f0', fg='#333333').grid(row=156, column=0, columnspan=2, pady=5)
        self.phase3_others_category_entry = create_label_entry(157, "Category")
        self.phase3_others_training_period_entry = create_label_entry(158, "Training Period")
        self.phase3_others_experiences_entry = create_label_entry(159, "Experiences")



        # Phase 4 Section
        tk.Label(self.scrollable_frame, text="Phase 4", font=section_title_font, fg=section_title_fg).grid(row=160, column=0, columnspan=2, pady=10)

        self.phase4_interview_no_entry = create_label_entry(161, "Interview No")
        self.phase4_family_status_entry = create_label_entry(162, "Family Status")
        self.phase4_mother_occupation_entry = create_label_entry(163, "Mother Occupation")
        self.phase4_father_occupation_entry = create_label_entry(164, "Father Occupation")
        self.phase4_family_class_entry = create_label_entry(165, "Family Class")
        self.phase4_family_position_entry = create_label_entry(166, "Family Position")

        stay_options = ["Yes", "Pending", "No"]
        self.phase4_stay_var = create_dropdown(167, "Possible to Stay With Us Minimum of 2 Years", stay_options)

        self.phase4_phone_number_entry = create_label_entry(168, "Phone Number")

        # Registered List Section
        tk.Label(self.scrollable_frame, text="Registered List", font=section_title_font, fg=section_title_fg).grid(row=169, column=0, columnspan=2, pady=10)

        self.reg_se_no_entry = create_label_entry(170, "Se.No")
        self.reg_in_no_entry = create_label_entry(171, "IN.No")
        self.reg_name_entry = create_label_entry(172, "Name")
        self.reg_id_number_entry = create_label_entry(173, "ID Number")
        self.reg_age_entry = create_label_entry(174, "Age")

        gender_options = ["Male", "Female"]
        self.reg_gender_var = create_dropdown(175, "Gender", gender_options)

        self.reg_gmail_entry = create_label_entry(176, "Gmail")
        self.reg_mobile_entry = create_label_entry(177, "Mobile")
        self.reg_mother_mobile_entry = create_label_entry(178, "Mother Mobile")
        self.reg_father_mobile_entry = create_label_entry(179, "Father Mobile")
        self.reg_whatsapp_entry = create_label_entry(180, "Whatsapp")
        self.reg_tracking_suit_entry = create_label_entry(181, "Tracking Suit")
        self.reg_paid_amount_entry = create_label_entry(182, "Paid Amount")
        self.reg_balance_entry = create_label_entry(183, "Balance")

        # Accept Checkbox
        self.accept_var = tk.IntVar()
        self.accept_checkbox = tk.Checkbutton(self.scrollable_frame, text="I Accept", variable=self.accept_var, font=entry_font, bg=label_bg)
        self.accept_checkbox.grid(row=184, column=0, columnspan=2, pady=5)


    # Submit Button
        self.submit_button = tk.Button(self.scrollable_frame, text="Submit", font=("Arial", 14, "bold"), bg="#4CAF50", fg="white", command=self.submit_data)
        self.submit_button.grid(row=185, column=0, columnspan=2, pady=10)

    def focus_next_widget(self, event):
        event.widget.tk_focusNext().focus()
        return "break"

    def submit_data(self):
        
        
        # Additional Fields
        full_name = self.full_name_entry.get()
        initial_name = self.initial_name_entry.get()
        dob = self.dob_entry.get()
        nic = self.nic_entry.get()
        gender = self.gender_var.get()
        postal_address = self.postal_address_entry.get()
        id_address = self.id_address_entry.get()
        permanent_address = self.permanent_address_entry.get()
        current_address = self.current_address_entry.get()
        mobile_number = self.mobile_number_entry.get()
        land_phone_number = self.land_phone_number_entry.get()
        whatsapp_number = self.whatsapp_number_entry.get()
        facebook_url = self.facebook_url_entry.get()
        instagram_id = self.instagram_id_entry.get()
        email = self.email_entry.get()

        province = self.province_var.get()
        district = self.district_var.get()
        ds_division = self.ds_division_var.get()
        gn_division = self.gn_division_entry.get()
        nearest_city = self.nearest_city_entry.get()

        guardian_name = self.guardian_name_entry.get()
        guardian_dob = self.guardian_dob_entry.get()
        guardian_age = self.guardian_age_entry.get()
        guardian_occupation = self.guardian_occupation_entry.get()
        guardian_contact_number = self.guardian_contact_number_entry.get()
        guardian_whatsapp_number = self.guardian_whatsapp_number_entry.get()
        guardian_address = self.guardian_address_entry.get()
        guardian_email = self.guardian_email_entry.get()

        ol_school = self.ol_school_entry.get()
        ol_admission_number = self.ol_admission_number_entry.get()
        ol_year = self.ol_year_entry.get()
        ol_award_achievements = self.ol_award_achievements_entry.get()

        al_school = self.al_school_entry.get()
        al_admission_number = self.al_admission_number_entry.get()
        al_year = self.al_year_entry.get()
        al_award_achievements = self.al_award_achievements_entry.get()

        pro_institution = self.pro_institution_entry.get()
        pro_result_achievement = self.pro_result_achievement_entry.get()
        pro_pass_out_year = self.pro_pass_out_year_entry.get()
        pro_areas_covered = self.pro_areas_covered_entry.get()
        pro_any_other = self.pro_any_other_entry.get()

        other_qualifications = self.other_qualifications_entry.get()

        special_talent = self.special_talent_entry.get()

        employment_institution = self.employment_institution_entry.get()
        employment_job_title = self.employment_job_title_entry.get()
        employment_joining_year = self.employment_joining_year_entry.get()
        employment_year_left = self.employment_year_left_entry.get()

        student_no = self.student_no_entry.get()
        student_name = self.student_name_entry.get()
        program_name = self.program_name_entry.get()
        jot_activity1 = self.jot_activity1_entry.get()
        phicycle_activity1 = self.phicycle_activity1_entry.get()
        lab_section1 = self.lab_section1_entry.get()
        assignment1 = self.assignment1_entry.get()
        assessment1 = self.assessment1_entry.get()
        practical1 = self.practical1_entry.get()
        module1_title = self.module1_title_entry.get()
        module1_subject1 = self.module1_subject1_entry.get()
        
        # Get values from Spinbox widgets
        module1_hours = f"{self.module1_hours_spinbox.get()}:{self.module1_minutes_spinbox.get()}"
        subject_hours = f"{self.subject_hours_spinbox.get()}:{self.subject_minutes_spinbox.get()}"
        
        module1_supervisor = self.module1_supervisor_entry.get()
        module1_lecturer = self.module1_lecturer_entry.get()
        module1_satisfaction = self.module1_satisfaction_entry.get()
        module1_lecturer_no = self.module1_lecturer_no_entry.get()
        module1_supervisor_no = self.module1_supervisor_no_entry.get()
        module1_title_no = self.module1_title_no_entry.get()
        module1_subject_no = self.module1_subject_no_entry.get()
        
        module_date = self.module_date_entry.get_date()
        subject_date = self.subject_date_entry.get_date()
        final_approval = self.final_approval_entry.get()

        # Get Phase 1 data
        se_no = self.se_no_entry.get()
        in_no = self.in_no_entry.get()
        phase1_name = self.phase1_name_entry.get()
        phase1_age = self.phase1_age_entry.get()
        phase1_phone = self.phase1_phone_entry.get()
        phase1_whatsapp = self.phase1_whatsapp_entry.get()
        phase1_email = self.phase1_email_entry.get()
        phase1_address = self.phase1_address_entry.get()
        phase1_ol_result = self.phase1_ol_result_entry.get()
        phase1_note = self.phase1_note_var.get()
        phase1_ol_sheet = self.phase1_ol_sheet_var.get()
        phase1_birth_cert = self.phase1_birth_cert_entry.get()
        phase1_id_copy = self.phase1_id_copy_entry.get()
        phase1_gs_copy = self.phase1_gs_copy_entry.get()
        phase1_edu_cert = self.phase1_edu_cert_entry.get()

        # Get Phase 2 data
        phase2_serial_no = self.phase2_serial_no_entry.get()
        phase2_interview_no = self.phase2_interview_no_entry.get()
        phase2_name = self.phase2_name_entry.get()
        phase2_height = self.phase2_height_entry.get()
        phase2_weight = self.phase2_weight_entry.get()
        phase2_blood_group = self.phase2_blood_group_entry.get()
        phase2_dress_code = self.phase2_dress_code_var.get()
        phase2_appearance = self.phase2_appearance_var.get()
        phase2_course_understanding = self.phase2_course_understanding_var.get()
        phase2_age = self.phase2_age_entry.get()
        phase2_location = self.phase2_location_entry.get()
        phase2_distance = self.phase2_distance_entry.get()
        phase2_status = self.phase2_status_var.get()

        phase3_interview_no = self.phase3_interview_no_entry.get()
        phase3_gce_ol = self.phase3_gce_ol_entry.get()
        phase3_gce_al = self.phase3_gce_al_entry.get()
        phase3_higher_study = self.phase3_higher_study_entry.get()
        phase3_train_by = self.phase3_train_by_entry.get()
        phase3_self = self.phase3_self_entry.get()
        phase3_voc_category = self.phase3_voc_category_entry.get()
        phase3_voc_training_period = self.phase3_voc_training_period_entry.get()
        phase3_voc_experiences = self.phase3_voc_experiences_entry.get()
        phase3_prof_category = self.phase3_prof_category_entry.get()
        phase3_prof_institute = self.phase3_prof_institute_entry.get()
        phase3_prof_experience = self.phase3_prof_experience_entry.get()
        phase3_others_category = self.phase3_others_category_entry.get()
        phase3_others_training_period = self.phase3_others_training_period_entry.get()
        phase3_others_experiences = self.phase3_others_experiences_entry.get()

         # Get Phase 4 data
        phase4_interview_no = self.phase4_interview_no_entry.get()
        phase4_family_status = self.phase4_family_status_entry.get()
        phase4_mother_occupation = self.phase4_mother_occupation_entry.get()
        phase4_father_occupation = self.phase4_father_occupation_entry.get()
        phase4_family_class = self.phase4_family_class_entry.get()
        phase4_family_position = self.phase4_family_position_entry.get()
        phase4_stay = self.phase4_stay_var.get()
        phase4_phone_number = self.phase4_phone_number_entry.get()

         # Get Registered List data
        reg_se_no = self.reg_se_no_entry.get()
        reg_in_no = self.reg_in_no_entry.get()
        reg_name = self.reg_name_entry.get()
        reg_id_number = self.reg_id_number_entry.get()
        reg_age = self.reg_age_entry.get()
        reg_gender = self.reg_gender_var.get()
        reg_gmail = self.reg_gmail_entry.get()
        reg_mobile = self.reg_mobile_entry.get()
        reg_mother_mobile = self.reg_mother_mobile_entry.get()
        reg_father_mobile = self.reg_father_mobile_entry.get()
        reg_whatsapp = self.reg_whatsapp_entry.get()
        reg_tracking_suit = self.reg_tracking_suit_entry.get()
        reg_paid_amount = self.reg_paid_amount_entry.get()
        reg_balance = self.reg_balance_entry.get()

        

        agree = self.agree_var.get()

        if agree != 1:
            messagebox.showwarning("Warning", "You must agree to the declaration.")
            return
        
        if self.accept_var.get() != 1:
            messagebox.showwarning("Warning", "You must accept before submitting.")
            return
        
         # Create a list of column headers
        headers = [
        "Full Name", "Initial Name", "Date of Birth", "NIC", "Gender", "Postal Address", "ID Address", 
        "Permanent Address", "Current Address", "Mobile Number", "Land Phone Number", "WhatsApp Number", 
        "Facebook URL", "Instagram ID", "Email", "Province", "District", "DS Division", "GN Division", 
        "Nearest City", "Guardian Name", "Guardian DOB", "Guardian Age", "Guardian Occupation", 
        "Guardian Contact Number", "Guardian WhatsApp Number", "Guardian Address", "Guardian Email", 
        "O/L School", "O/L Admission Number", "O/L Year", "O/L Award Achievements", "A/L School", 
        "A/L Admission Number", "A/L Year", "A/L Award Achievements", "Professional Institution", 
        "Professional Result Achievement", "Professional Pass Out Year", "Professional Areas Covered", 
        "Professional Any Other", "Other Qualifications", "Special Talent", "Employment Institution", 
        "Employment Job Title", "Employment Joining Year", "Employment Year Left", "Student No", 
        "Student Name", "Program Name", "JOT Activity", "Phicycle Activity", "Lab Section", "Assignment", 
        "Assessment", "Practical", "Module Title", "Module Subject", "Module Hours", "Subject Hours", 
        "Module Supervisor", "Module Lecturer", "Module Satisfaction", "Module Lecturer No", 
        "Module Supervisor No", "Module Title No", "Module Subject No", "Module Date", "Subject Date", 
        "Final Approval", "SE No", "IN No", "Phase 1 Name", "Phase 1 Age", "Phase 1 Phone", 
        "Phase 1 WhatsApp", "Phase 1 Email", "Phase 1 Address", "Phase 1 O/L Result", "Phase 1 Note", 
        "Phase 1 O/L Sheet", "Phase 1 Birth Certificate", "Phase 1 ID Copy", "Phase 1 GS Copy", 
        "Phase 1 Education Certificate", "Phase 2 Serial No", "Phase 2 Interview No", "Phase 2 Name", 
        "Phase 2 Height", "Phase 2 Weight", "Phase 2 Blood Group", "Phase 2 Dress Code", 
        "Phase 2 Appearance", "Phase 2 Course Understanding", "Phase 2 Age", "Phase 2 Location", 
        "Phase 2 Distance", "Phase 2 Status", "Phase 3 Interview No", "Phase 3 GCE O/L", "Phase 3 GCE A/L", 
        "Phase 3 Higher Study", "Phase 3 Train By", "Phase 3 Self", "Phase 3 Vocational Category", 
        "Phase 3 Vocational Training Period", "Phase 3 Vocational Experiences", 
        "Phase 3 Professional Category", "Phase 3 Professional Institute", 
        "Phase 3 Professional Experience", "Phase 3 Others Category", 
        "Phase 3 Others Training Period", "Phase 3 Others Experiences", "Phase 4 Interview No", 
        "Phase 4 Family Status", "Phase 4 Mother Occupation", "Phase 4 Father Occupation", 
        "Phase 4 Family Class", "Phase 4 Family Position", "Phase 4 Stay", "Phase 4 Phone Number", 
        "Reg SE No", "Reg IN No", "Reg Name", "Reg ID Number", "Reg Age", "Reg Gender", "Reg Gmail", 
        "Reg Mobile", "Reg Mother Mobile", "Reg Father Mobile", "Reg WhatsApp", "Reg Tracking Suit", 
        "Reg Paid Amount", "Reg Balance"
    ]

        

        # Save the data into an Excel file
        try:
            file_path = "Kanishka_PROGRAMME.xlsx"

            # Check if the file exists and is empty
            file_exists = os.path.exists(file_path)
            file_empty = False
            if file_exists:



              wb = load_workbook(file_path)
              ws = wb.active
              if ws.max_row == 1 and ws.max_column == 1 and ws.cell(row=1, column=1).value is None:
                file_empty = True

            else:
                 wb = Workbook()
                 ws = wb.active
                 file_empty = True

             # If file doesn't exist or is empty, write headers
            if not file_exists or file_empty:
                ws.append(headers)




            # Write the data to the Excel sheet
            ws.append([
                full_name, initial_name, dob, nic, gender, postal_address, id_address, permanent_address,
                current_address, mobile_number, land_phone_number, whatsapp_number, facebook_url, instagram_id, email, province, 
                district, ds_division, gn_division, nearest_city, guardian_name, guardian_dob, guardian_age, guardian_occupation, 
                guardian_contact_number, guardian_whatsapp_number, guardian_address, guardian_email, ol_school, ol_admission_number, 
                ol_year, ol_award_achievements, al_school, al_admission_number, al_year, al_award_achievements, pro_institution, 
                pro_result_achievement, pro_pass_out_year, pro_areas_covered, pro_any_other, other_qualifications, special_talent, 
                employment_institution, employment_job_title, employment_joining_year, employment_year_left , student_no, student_name, program_name, jot_activity1, phicycle_activity1, lab_section1, assignment1, assessment1,
                practical1, module1_title, module1_subject1, module1_hours, subject_hours, module1_supervisor, module1_lecturer,
                module1_satisfaction, module1_lecturer_no, module1_supervisor_no, module1_title_no, module1_subject_no, module_date,
                subject_date, final_approval, se_no, in_no, phase1_name, phase1_age, phase1_phone, phase1_whatsapp, phase1_email, phase1_address,
                phase1_ol_result, phase1_note, phase1_ol_sheet, phase1_birth_cert, phase1_id_copy, phase1_gs_copy,
                phase1_edu_cert, phase2_serial_no, phase2_interview_no, phase2_name, phase2_height, phase2_weight,
                phase2_blood_group, phase2_dress_code, phase2_appearance, phase2_course_understanding, phase2_age,
                phase2_location, phase2_distance, phase2_status, phase3_interview_no, phase3_gce_ol, phase3_gce_al, phase3_higher_study,
                phase3_train_by, phase3_self, phase3_voc_category, phase3_voc_training_period,
                phase3_voc_experiences, phase3_prof_category, phase3_prof_institute,
                phase3_prof_experience, phase3_others_category, phase3_others_training_period,
                phase3_others_experiences, phase4_interview_no, phase4_family_status, phase4_mother_occupation,
                phase4_father_occupation, phase4_family_class, phase4_family_position,
                phase4_stay, phase4_phone_number, reg_se_no, reg_in_no, reg_name,
                reg_id_number, reg_age, reg_gender, reg_gmail, reg_mobile,
                reg_mother_mobile, reg_father_mobile, reg_whatsapp, reg_tracking_suit,
                reg_paid_amount, reg_balance

        # Add fields for Phase 3, Phase 4, Registered List, and Board Members when they are defined
            ])

            # Save the workbook
            wb.save(file_path)

            messagebox.showinfo("Success", "Data saved successfully.")

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while saving the data: {str(e)}")

    def new_method(self):
        df = pd.DataFrame()
        return df

# Main application loop
if __name__ == "__main__":
    root = tk.Tk()
    app = StudentDataApp(root)
    root.mainloop()
